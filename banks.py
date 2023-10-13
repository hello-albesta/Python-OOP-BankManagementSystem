import os.path
import time as system_time

import pandas as pd
from oauth2client.service_account import ServiceAccountCredentials as Credentials
import gspread
from openpyxl import load_workbook as workbook
from tabulate import tabulate

from custom_defined import const_dict as const


class Bank:
    def __init__(self,
                 accounts):
        self.accounts = accounts

    """
    Retrieves the selected user account data.

    Returns:
        pandas.DataFrame: The user account data for the selected account.

    ==================================================================================================================

    The explanation of the code snippet:

    This code snippet defines a method called get_selected_user_account_data that reads an Excel file and
    filters the data based on the values of ACC_NAME_KEY and ACC_PASS_KEY in the user_account_list DataFrame.

    It returns the filtered data.
    """
    def get_selected_user_account_data(self):
        user_account_list = pd.read_excel(const.LOCAL_FOLDER_EXPORT_FILE_PATH + const.EXCEL_USER_ACC_LIST)

        get_user_account_data = user_account_list[(user_account_list[const.ACC_NAME_KEY] ==
                                                  self.accounts[const.ACC_NAME_KEY])
                                                  & (user_account_list[const.ACC_PASS_KEY] ==
                                                  self.accounts[const.ACC_PASS_KEY])]

        return get_user_account_data

    """
    Retrieves the transaction history data for the selected user.

    Returns:
        pandas.DataFrame: The transaction history data for the selected user.
        0: If the local export file for user transaction history does not exist.

    ==================================================================================================================

    The explanation of the code snippet:

    This code defines a function called get_selected_user_transact_hist_data that retrieves transaction history data
    for a selected user.

    It first constructs the path to the local export file where the transaction history data is stored.

    Then, it checks if the file exists.
    If it does, it reads the data from the file using the pd.read_excel function from the pandas library.

    Next, it filters the data to only include rows where the account number matches
    the selected user's account number.

    Finally, it returns the filtered data. If the file does not exist, it returns 0.
    """
    def get_selected_user_transact_hist_data(self):
        local_export_user_transact_hist_data = (const.LOCAL_FOLDER_EXPORT_FILE_PATH +
                                                const.EXCEL_USER_ACC_TRANSACTION_HISTORY)

        if os.path.exists(local_export_user_transact_hist_data):
            user_account_transact_hist = pd.read_excel(local_export_user_transact_hist_data)

            get_user_transact_hist_data = user_account_transact_hist[
                user_account_transact_hist[const.ACC_NUMBER_KEY].str.contains(
                    self.accounts[const.ACC_NUMBER_KEY])
            ]

            return get_user_transact_hist_data

        else:
            return 0

    """
    Displays the transaction history for the selected user.

    Parameters:
        None

    Returns:
        None

    ==================================================================================================================

    The explanation of the code snippet:

    This code defines a method called display_transact_hist that prints the transaction history of a user.

    It first prints the name of the user's account.
    Then, it calls a method called get_selected_user_transact_hist_data to retrieve the transaction history data.

    If the returned data is a non-empty DataFrame,
    it prints the data in a formatted table using the tabulate function.

    If the data is empty or not a DataFrame, it prints a message indicating that no data is available.
    """
    def display_transact_hist(self):
        print(f"{self.accounts[const.ACC_NAME_KEY]}'s TRANSACTION HISTORY:")

        get_user_transact_hist_data = self.get_selected_user_transact_hist_data()

        if isinstance(get_user_transact_hist_data, pd.DataFrame):
            if len(get_user_transact_hist_data) > 0:
                print("<< USER TRANSACTION HISTORY INFORMATION TABLE >>")
                print(tabulate(tabular_data=get_user_transact_hist_data,
                               headers='keys',
                               tablefmt='fancy_grid',
                               showindex=False,
                               stralign="center",
                               numalign="right"))
            else:
                print(const.NO_DATA_AVAILABLE)

        else:
            print(const.NO_DATA_AVAILABLE)

    """
    Exports the transaction history to a Google Spreadsheet.

    Args:
        acc_transact_hist (dict): The transaction history for the account.

    Returns:
        None

    Raises:
        None

    ==================================================================================================================

    The explanation of the code snippet:

    This code snippet defines a function gdoc_online_export_transact_hist that exports transaction history data to
    a Google Spreadsheet.

    The function takes two parameters: self (which refers to an object instance) and
    acc_transact_hist (which is a dictionary containing transaction history data).

    The function creates a dictionary gspreed_transact_hist by extracting specific values
    from acc_transact_hist and self.

    It then appends the values of gspreed_transact_hist to an empty list spreed_transact_hist.

    The function uses Google Sheets API to authenticate with the Google Sheets service and open a spreadsheet using
    the provided credentials.

    Based on the type of savings account (self.accounts[const.ACC_SAVING_TYPE_KEY]),
    the function appends the names of additional sheets to the excel_spreadsheet_list and
    creates a string user_acc_spreadsheet_name with the names of the transaction sheet and the additional sheets.

    The function iterates over the excel_spreadsheet_list and
    performs different operations depending on the current sheet.

    For the transaction sheet, it appends the values in spreed_transact_hist to the sheet,
    retrieves all records from the sheet, and filters the records based on certain conditions.

    For the additional sheets, it only appends the values in spreed_transact_hist to the sheet.

    Finally, the function prints a success message indicating that the Google Spreadsheet has been updated with
    the transaction data.
    """
    def gdoc_online_export_transact_hist(self, acc_transact_hist):
        gspreed_transact_hist = {
            const.ACC_DATE_KEY: acc_transact_hist[const.ACC_DATE_KEY],
            const.ACC_NUMBER_KEY: self.accounts[const.ACC_NUMBER_KEY],
            const.ACC_NAME_KEY: self.accounts[const.ACC_NAME_KEY],
            const.ACC_BALANCE_KEY: acc_transact_hist[const.ACC_BALANCE_KEY],
            const.ACC_STUDENT_KEY: const.CREATOR_NAME
        }

        spreed_transact_hist = []
        spreed_transact_hist[len(spreed_transact_hist):] = [i for i in gspreed_transact_hist.values()]

        creds = Credentials.from_json_keyfile_name(filename=const.CREDENTIALS_FILE_NAME_PATH,
                                                   scopes=const.CREDENTIALS_SCOPE_URl)
        client = gspread.authorize(credentials=creds)
        spreadsheet = client.open_by_key(const.CREDENTIALS_SHEET_KEY)

        excel_spreadsheet_list = [const.EXPORT_EXCEL_FILE_TRANSACTION_SHEET]

        if self.accounts[const.ACC_SAVING_TYPE_KEY] == const.ACC_CONVENTIONAL_SAVING_TYPE:
            excel_spreadsheet_list.append(const.EXPORT_EXCEL_FILE_CONVENTIONAL_SHEET)

            user_acc_spreadsheet_name = f"{const.EXPORT_EXCEL_FILE_TRANSACTION_SHEET}, " \
                                        f"{const.EXPORT_EXCEL_FILE_CONVENTIONAL_SHEET}"

        else:
            excel_spreadsheet_list.append(const.EXPORT_EXCEL_FILE_SHARIA_SHEET)

            user_acc_spreadsheet_name = f"{const.EXPORT_EXCEL_FILE_TRANSACTION_SHEET}, " \
                                        f"{const.EXPORT_EXCEL_FILE_SHARIA_SHEET}"

        for excel_spreadsheet in excel_spreadsheet_list:
            if excel_spreadsheet == const.EXPORT_EXCEL_FILE_TRANSACTION_SHEET:
                (spreadsheet.worksheet(const.EXPORT_EXCEL_FILE_TRANSACTION_SHEET)
                 .append_row(values=spreed_transact_hist,
                             value_input_option=const.EXPORT_EXCEL_VALUE_INPUT_OPTION))
                transaction_sheet = spreadsheet.worksheet(const.EXPORT_EXCEL_FILE_TRANSACTION_SHEET).get_all_records()

                user_account_transact_list = []

                for transaction_hist in transaction_sheet:
                    if (transaction_hist[const.ACC_STUDENT_KEY] == const.CREATOR_NAME and
                            transaction_hist[const.ACC_NUMBER_KEY] == self.accounts[const.ACC_NUMBER_KEY] and
                            transaction_hist[const.ACC_NAME_KEY] == self.accounts[const.ACC_NAME_KEY]):

                        user_account_transact_hist_filter = {
                            const.ACC_SPREADSHEET_KEY: user_acc_spreadsheet_name,
                            const.ACC_DATE_KEY: transaction_hist[const.ACC_DATE_KEY],
                            const.ACC_NUMBER_KEY: transaction_hist[const.ACC_NUMBER_KEY],
                            const.ACC_NAME_KEY: transaction_hist[const.ACC_NAME_KEY],
                            const.ACC_BALANCE_KEY: transaction_hist[const.ACC_BALANCE_KEY]
                        }

                        user_account_transact_list.append(user_account_transact_hist_filter)
                print(f"======================================================================================\n"
                      f"{const.GDOC_FILE_UPDATED_SUCCESSFULLY} ::> {spreadsheet.title}\n"
                      f"ON SPREADSHEET ::> [{const.EXPORT_EXCEL_FILE_TRANSACTION_SHEET}]\n")

            elif excel_spreadsheet == const.EXPORT_EXCEL_FILE_CONVENTIONAL_SHEET:
                (spreadsheet.worksheet(const.EXPORT_EXCEL_FILE_CONVENTIONAL_SHEET)
                 .append_row(values=spreed_transact_hist,
                             value_input_option=const.EXPORT_EXCEL_VALUE_INPUT_OPTION))
                print(f"{const.GDOC_FILE_UPDATED_SUCCESSFULLY} ::> {spreadsheet.title}\n"
                      f"ON SPREADSHEET ::> [{const.EXPORT_EXCEL_FILE_CONVENTIONAL_SHEET}]\n"
                      f"======================================================================================")
                print("\n")

            elif excel_spreadsheet == const.EXPORT_EXCEL_FILE_SHARIA_SHEET:
                (spreadsheet.worksheet(const.EXPORT_EXCEL_FILE_SHARIA_SHEET)
                 .append_row(values=spreed_transact_hist,
                             value_input_option=const.EXPORT_EXCEL_VALUE_INPUT_OPTION))

                print(f"{const.GDOC_FILE_UPDATED_SUCCESSFULLY} ::> {spreadsheet.title}\n"
                      f"ON SPREADSHEET ::> [{const.EXPORT_EXCEL_FILE_SHARIA_SHEET}]\n"
                      f"======================================================================================")
                print("\n")

    """
    Generates a function comment for the given function body.

    Args:
        acc_transact_hist (dict): The account transaction history.

    Returns:
        bool: True if the function execution is successful, False otherwise.

    ==================================================================================================================

    The explanation of the code snippet:

    This code snippet defines a function excel_offline_new_export_transact_hist that takes in two arguments:
    self and acc_transact_hist.

    Inside the function, a dictionary account is created by extracting certain values from acc_transact_hist and
    self.accounts.

    Based on the value of acc_transact_hist["transact_hist"],
    a pandas DataFrame df_transact_hist is created either from a single account or from the account dictionary.

    An excel_spreadsheet_list is created with the value [const.EXPORT_EXCEL_FILE_TRANSACTION_SHEET].
    If the value of account[const.ACC_SAVING_TYPE_KEY] is const.ACC_CONVENTIONAL_SAVING_TYPE,
    then const.EXPORT_EXCEL_FILE_CONVENTIONAL_SHEET is appended to the excel_spreadsheet_list,
    otherwise const.EXPORT_EXCEL_FILE_SHARIA_SHEET is appended.

    A local export file path is defined based on constants and if the path does not exist,
    a new Excel file is created using pd.ExcelWriter.
    The df_transact_hist DataFrame is written to the Excel file for each spreadsheet in excel_spreadsheet_list.

    If the local export file path already exists, the code tries to open it as a read-only workbook.
    Then, using the same pd.ExcelWriter, the df_transact_hist DataFrame is appended to the existing Excel file
    for each spreadsheet in excel_spreadsheet_list.

    There are exception handlers for PermissionError and FileNotFoundError which print error messages and
    return False if these exceptions occur.

    The function returns True if the code successfully creates or updates the Excel file, and False otherwise.
    """
    def excel_offline_new_export_transact_hist(self, acc_transact_hist):
        account = {
            const.ACC_DATE_KEY: acc_transact_hist[const.ACC_DATE_KEY],
            const.ACC_SAVING_TYPE_KEY: self.accounts[const.ACC_SAVING_TYPE_KEY],
            const.ACC_NUMBER_KEY: self.accounts[const.ACC_NUMBER_KEY],
            const.ACC_NAME_KEY: self.accounts[const.ACC_NAME_KEY],
            const.ACC_BALANCE_KEY: self.accounts[const.ACC_BALANCE_KEY],
            const.ACC_TRANSACT_HIST_KEY: acc_transact_hist[const.ACC_TRANSACT_HIST_KEY],
        }

        if acc_transact_hist["transact_hist"] == const.DEF_ACC_TRANSACT_HIST_DATA:
            df_transact_hist = pd.DataFrame([account])

        else:
            df_transact_hist = pd.DataFrame(account)

        excel_spreadsheet_list = [const.EXPORT_EXCEL_FILE_TRANSACTION_SHEET]

        if account[const.ACC_SAVING_TYPE_KEY] == const.ACC_CONVENTIONAL_SAVING_TYPE:
            excel_spreadsheet_list.append(const.EXPORT_EXCEL_FILE_CONVENTIONAL_SHEET)

        else:
            excel_spreadsheet_list.append(const.EXPORT_EXCEL_FILE_SHARIA_SHEET)

        local_export_file_path = const.LOCAL_FOLDER_EXPORT_FILE_PATH + const.EXCEL_USER_ACC_TRANSACTION_HISTORY

        if not os.path.exists(local_export_file_path):
            with pd.ExcelWriter(path=local_export_file_path,
                                engine=const.EXCEl_FILE_ENGINE,
                                mode="w") as excel_new_writer:
                for excel_spreadsheet in excel_spreadsheet_list:
                    df_transact_hist.to_excel(excel_writer=excel_new_writer,
                                              index=False,
                                              sheet_name=excel_spreadsheet)

                    print(f"======================================================================================\n"
                          f"{const.EXCEL_FILE_CREATED_SUCCESSFULLY} ::> {const.EXCEL_USER_ACC_TRANSACTION_HISTORY}\n"
                          f"ON SPREADSHEET ::> [{excel_spreadsheet}]\n"
                          f"======================================================================================")

            return True

        else:
            try:
                excel_local_file = workbook(local_export_file_path, read_only=True)

                with pd.ExcelWriter(path=local_export_file_path,
                                    engine=const.EXCEl_FILE_ENGINE,
                                    mode="a",
                                    if_sheet_exists="overlay") as excel_exist_updater:
                    for excel_spreadsheet in excel_spreadsheet_list:
                        if excel_spreadsheet in excel_local_file.sheetnames:
                            df_transact_hist.to_excel(excel_writer=excel_exist_updater,
                                                      index=False,
                                                      header=False,
                                                      sheet_name=excel_spreadsheet,
                                                      startrow=excel_exist_updater
                                                      .sheets[excel_spreadsheet].max_row)

                        else:
                            df_transact_hist.to_excel(excel_writer=excel_exist_updater,
                                                      index=False,
                                                      sheet_name=excel_spreadsheet)

                        print(f"================================================================================="
                              f"=====\n"
                              f"{const.EXCEL_FILE_UPDATED_SUCCESSFULLY} ::> "
                              f"{const.EXCEL_USER_ACC_TRANSACTION_HISTORY}\n"
                              f"ON SPREADSHEET ::> [{excel_spreadsheet}]\n"
                              f"=================================================================================="
                              f"====")

                return True

            except PermissionError:
                system_time.sleep(1)

                print(const.FILE_PERMISSION_ERROR_HANDLER_ERROR)
                print(
                    f"{const.WARNING_TEXT_1}Maybe the file {const.EXCEL_USER_ACC_TRANSACTION_HISTORY} "
                    f"is being opened by you or others,\n"
                    f"please close the file {const.EXCEL_USER_ACC_TRANSACTION_HISTORY} "
                    f"first and then try to transaction again!{const.WARNING_TEXT_2}")
                print("\n")
                print(f"{const.WARNING_TEXT_1}If the issue still persists, please do not hesitate to contact me,\n"
                      f"({const.CREATOR_NAME})!{const.WARNING_TEXT_2}")
                print("\n")

                return False

            except FileNotFoundError:
                system_time.sleep(1)

                print(const.FILE_PERMISSION_NOT_FOUND_HANDLER_ERROR)
                print(f"{const.WARNING_TEXT_1}File {const.EXCEL_USER_ACC_TRANSACTION_HISTORY} was not found!,\n"
                      f"make sure the file {const.EXCEL_USER_ACC_TRANSACTION_HISTORY} "
                      f"is correct and then try to transaction again!{const.WARNING_TEXT_2}")
                print("\n")
                print(f"{const.WARNING_TEXT_1}If the issue still persists, please do not hesitate to contact me,\n"
                      f"({const.CREATOR_NAME})!{const.WARNING_TEXT_2}")
                print("\n")

                return False

    """
    Export user account history to an offline Excel file.

    :return: True if the export is successful, False otherwise.

    ==================================================================================================================

    The explanation of the code snippet:

    This code snippet is a method called excel_offline_export_user_acc_hist
    that exports user account history data to an Excel file.

    It takes the user's account information, creates a DataFrame using pandas,
    and saves it to a specified file path.

    If the file does not exist, a new Excel file is created and the DataFrame is written to it.

    If the file already exists, the DataFrame is appended to the existing sheet.
    The code also includes error handling for cases where there are permission errors or the file is not found.
    """
    def excel_offline_export_user_acc_hist(self):
        user_account_export_format = {
            const.ACC_NUMBER_KEY: self.accounts[const.ACC_NUMBER_KEY],
            const.ACC_NAME_KEY: self.accounts[const.ACC_NAME_KEY],
            const.ACC_PASS_KEY: self.accounts[const.ACC_PASS_KEY],
            const.ACC_SAVING_TYPE_KEY: self.accounts[const.ACC_SAVING_TYPE_KEY],
        }

        df_user_acc_hist = pd.DataFrame([user_account_export_format])

        local_export_file_path = const.LOCAL_FOLDER_EXPORT_FILE_PATH + const.EXCEL_USER_ACC_LIST

        if not os.path.isdir(const.LOCAL_FOLDER_EXPORT_FILE_PATH):
            temp_local_folder_export_dir = const.LOCAL_FOLDER_EXPORT_FILE_PATH
            temp_local_folder_export_dir_path = os.path.join(os.getcwd() + '/', temp_local_folder_export_dir)

            os.mkdir(temp_local_folder_export_dir_path)

        if not os.path.exists(local_export_file_path):
            with pd.ExcelWriter(path=local_export_file_path,
                                engine=const.EXCEl_FILE_ENGINE,
                                mode="w") as excel_new_writer:
                df_user_acc_hist.to_excel(excel_writer=excel_new_writer,
                                          index=False,
                                          sheet_name=const.EXPORT_EXCEL_USER_ACCOUNT_SHEET)

                print("<< USER ACCOUNT INFORMATION TABLE >>")
                print(tabulate(tabular_data=df_user_acc_hist,
                               headers='keys',
                               tablefmt='fancy_grid',
                               showindex=False,
                               stralign="center",
                               numalign="right"))
                print("\n")
                print(f"======================================================================================\n"
                      f"{const.EXCEL_FILE_CREATED_SUCCESSFULLY} ::> {const.EXCEL_USER_ACC_LIST}\n"
                      f"ON SPREADSHEET ::> [{const.EXPORT_EXCEL_USER_ACCOUNT_SHEET}]\n"
                      f"======================================================================================")
                print("\n")

            return True

        else:
            try:
                with pd.ExcelWriter(path=local_export_file_path,
                                    engine=const.EXCEl_FILE_ENGINE,
                                    mode="a",
                                    if_sheet_exists="overlay") as excel_exist_updater:
                    df_user_acc_hist.to_excel(excel_writer=excel_exist_updater,
                                              index=False,
                                              header=False,
                                              sheet_name=const.EXPORT_EXCEL_USER_ACCOUNT_SHEET,
                                              startrow=excel_exist_updater
                                              .sheets[const.EXPORT_EXCEL_USER_ACCOUNT_SHEET].max_row)

                    print("<< USER ACCOUNT INFORMATION TABLE >>")
                    print(tabulate(tabular_data=df_user_acc_hist,
                                   headers='keys',
                                   tablefmt='fancy_grid',
                                   showindex=False,
                                   stralign="center",
                                   numalign="right"))
                    print("\n")
                    print(f"======================================================================================\n"
                          f"{const.EXCEL_FILE_UPDATED_SUCCESSFULLY} ::> {const.EXCEL_USER_ACC_LIST}\n"
                          f"ON SPREADSHEET ::> [{const.EXPORT_EXCEL_USER_ACCOUNT_SHEET}]\n"
                          f"======================================================================================")
                    print("\n")

                    return True

            except PermissionError:
                system_time.sleep(1)

                print(const.FILE_PERMISSION_ERROR_HANDLER_ERROR)
                print(
                    f"{const.WARNING_TEXT_1}Maybe the file {const.EXCEL_USER_ACC_LIST} is being opened by "
                    f"you or others,\n"
                    f"please close the file {const.EXCEL_USER_ACC_LIST} first and then try to register again!"
                    f"{const.WARNING_TEXT_2}")
                print("\n")
                print(f"{const.WARNING_TEXT_1}If the issue still persists, please do not hesitate to contact me,\n"
                      f"({const.CREATOR_NAME})!{const.WARNING_TEXT_2}")
                print("\n")

                return False

            except FileNotFoundError:
                system_time.sleep(1)

                print(const.FILE_PERMISSION_NOT_FOUND_HANDLER_ERROR)
                print(f"{const.WARNING_TEXT_1}File {const.EXCEL_USER_ACC_LIST} was not found!,\n"
                      f"make sure the file {const.EXCEL_USER_ACC_LIST} is correct and then try to register again!"
                      f"{const.WARNING_TEXT_2}")
                print("\n")
                print(f"{const.WARNING_TEXT_1}If the issue still persists, please do not hesitate to contact me,\n"
                      f"({const.CREATOR_NAME})!{const.WARNING_TEXT_2}")
                print("\n")

                return False

    """
    Check if the data is registered.

    Returns:
        bool: True if the data is registered, False otherwise.

    ==================================================================================================================

    The explanation of the code snippet:

    This code snippet defines a function called check_registered_data that checks if a local export file exists.

    If the file does not exist, it returns False.

    If the file exists, it calls a method get_selected_user_account_data and prints the returned data.

    If the length of the returned data is greater than 1, it returns a constant value representing an error.

    If the length is greater than 0 but not greater than 1, it returns the data.

    If the length is 0, it also returns the constant value representing an error.
    """
    def check_registered_data(self):
        local_export_file_path = const.LOCAL_FOLDER_EXPORT_FILE_PATH + const.EXCEL_USER_ACC_LIST

        if not os.path.exists(local_export_file_path):
            return False

        get_user_account_data = self.get_selected_user_account_data()

        print(get_user_account_data)

        if len(get_user_account_data) > 1:
            return const.USER_ACC_NAME_AND_PASS_NOT_FOUND_ERROR

        elif len(get_user_account_data) > 0:
            return get_user_account_data

        else:
            return const.USER_ACC_NAME_AND_PASS_NOT_FOUND_ERROR

    """
    This function processes registered data.

    Parameters:
        None

    Returns:
        bool: True if the data was successfully processed, False otherwise.

    ==================================================================================================================

    The explanation of the code snippet:

    This code snippet defines a function called processed_registered_data that performs some operations on
    user account data.

    It checks if a local export file exists, retrieves data from the file, and performs some calculations and
    validations based on the data.

    It also prints some messages and returns True or False based on the success of exporting user account history
    to an Excel file.
    """
    def processed_registered_data(self):
        total_user_account = []

        local_export_file_path = const.LOCAL_FOLDER_EXPORT_FILE_PATH + const.EXCEL_USER_ACC_LIST

        if os.path.exists(local_export_file_path):
            check_exist_user_account_password_data = self.get_selected_user_account_data()

            print(check_exist_user_account_password_data)

            if len(check_exist_user_account_password_data) > 0:
                system_time.sleep(1)

                print(const.WARNING_TEXT_1 + const.USER_ACC_EXISTS_ERROR + const.WARNING_TEXT_2)

                system_time.sleep(1)

                print(const.PAGING_MENU_JUMP)

                return False

            user_account_list = pd.read_excel(const.LOCAL_FOLDER_EXPORT_FILE_PATH + const.EXCEL_USER_ACC_LIST)

            if self.accounts[const.ACC_SAVING_TYPE_KEY] == const.ACC_CONVENTIONAL_SAVING_TYPE:
                total_user_account = [conventional for conventional in user_account_list[const.ACC_NUMBER_KEY]
                                      if conventional[0] == const.ACC_CONVENTIONAL_SAVING_ID]

                print(f"total_user_account: {total_user_account}")

            else:
                total_user_account = [sharia for sharia in user_account_list[const.ACC_NUMBER_KEY]
                                      if sharia[0] == const.ACC_SHARIA_SAVING_ID]

                print(f"total_user_account: {total_user_account}")

        str_total_user_account = str(len(total_user_account) + const.DEF_ACC_NUMBER_LEADING_ZEROS) \
            if str(len(total_user_account)) != "0" \
            else str(const.DEF_ACC_NUMBER_LEADING_ZEROS)

        print(f"str_total_user_account: {str_total_user_account}")

        acc_number_leading_zeros = (len(str_total_user_account) + const.DEF_ACC_NUMBER_LEADING_ZEROS)

        if self.accounts[const.ACC_SAVING_TYPE_KEY] == const.ACC_CONVENTIONAL_SAVING_TYPE:
            self.accounts[const.ACC_NUMBER_KEY] = (const.ACC_CONVENTIONAL_SAVING_ID +
                                                   (str_total_user_account.zfill(acc_number_leading_zeros)))

        else:
            self.accounts[const.ACC_NUMBER_KEY] = (const.ACC_SHARIA_SAVING_ID +
                                                   (str_total_user_account.zfill(acc_number_leading_zeros)))

        res_excel_offline_export_user_acc_hist = self.excel_offline_export_user_acc_hist()

        if res_excel_offline_export_user_acc_hist:
            return True

        else:
            return False
